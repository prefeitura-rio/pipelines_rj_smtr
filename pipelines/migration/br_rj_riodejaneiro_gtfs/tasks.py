# -*- coding: utf-8 -*-
"""
Tasks for gtfs
"""
import io
import zipfile
from datetime import datetime
from functools import partial

import openpyxl as xl
import pandas as pd
from prefect import task
from prefeitura_rio.pipelines_utils.logging import log
from prefeitura_rio.pipelines_utils.redis_pal import get_redis_client

from pipelines.constants import constants
from pipelines.migration.br_rj_riodejaneiro_gtfs.utils import (
    download_controle_os_csv,
    download_file,
    download_xlsx,
    filter_valid_rows,
    processa_ordem_servico,
    processa_ordem_servico_faixa_horaria,
    processa_ordem_servico_trajeto_alternativo,
)
from pipelines.migration.utils import (
    create_bq_external_table,
    get_upload_storage_blob,
    save_raw_local_func,
)
from pipelines.utils.extractors.gdrive import get_google_api_service
from pipelines.utils.gcp.bigquery import BQTable
from pipelines.utils.gcp.storage import Storage


@task
def get_last_capture_os(dataset_id: str, mode: str = "prod") -> dict:
    """
    Retrieves the last captured OS for a given dataset ID and mode.

    Args:
        dataset_id (str): The ID of the dataset.
        mode (str, optional): The mode of operation. Defaults to "prod".

    Returns:
        dict: The last captured OS.

    """
    redis_client = get_redis_client()
    fetch_key = f"{dataset_id}.last_captured_os"
    if mode != "prod":
        fetch_key = f"{mode}.{fetch_key}"

    last_captured_os = redis_client.get(fetch_key)
    if last_captured_os is not None:
        last_captured_os = last_captured_os["last_captured_os"]

    #  verifica se last_capture_os tem formado dia/mes/ano_index
    if last_captured_os is not None:
        if "/" in last_captured_os:
            index = last_captured_os.split("_")[1]
            data = datetime.strptime(last_captured_os.split("_")[0], "%d/%m/%Y").strftime(
                "%Y-%m-%d"
            )
            last_captured_os = data + "_" + index

    log(f"Last captured os: {last_captured_os}")

    return last_captured_os


@task
def update_last_captured_os(dataset_id: str, data_index: str, mode: str = "prod") -> None:
    """
    Update the last captured operating system for a given dataset.

    Args:
        dataset_id (str): The ID of the dataset.
        data_index (str): The last captured operating system.
        mode (str, optional): The mode of operation. Defaults to "prod".

    Returns:
        None
    """
    redis_client = get_redis_client()
    fetch_key = f"{dataset_id}.last_captured_os"
    if mode != "prod":
        fetch_key = f"{mode}.{fetch_key}"
    last_captured_os = redis_client.get(fetch_key)
    #  verifica se last_capture_os tem formado dia/mes/ano_index e converte para ano-mes-dia_index
    if last_captured_os is not None:
        if "/" in last_captured_os:
            index = last_captured_os.split("_")[1]
            data = datetime.strptime(last_captured_os.split("_")[0], "%d/%m/%Y").strftime(
                "%Y-%m-%d"
            )
            last_captured_os = data + "_" + index
    # verifica se a ultima os capturada é maior que a nova
    if last_captured_os is not None:
        if last_captured_os["last_captured_os"] > data_index:
            return
    redis_client.set(fetch_key, {"last_captured_os": data_index})


@task(nout=4)
def get_os_info(last_captured_os: str = None, data_versao_gtfs: str = None) -> dict:
    """
    Retrieves information about the OS.

    Args:
        last_captured_os (str): The last captured OS data_index.

    Returns:
        tuple: A tuple containing the following elements:
            - flag_new_os (bool): Indicates whether a new OS was found.
            - data (dict): A dictionary containing the OS information.
            - data_index (str): The index of the captured OS.
            - inicio_vigencia_os (str): The start date of the captured OS.

    """
    df = download_controle_os_csv(constants.GTFS_CONTROLE_OS_URL.value)

    flag_new_os = False
    data = {"Início da Vigência da OS": None, "data_index": None}

    if df.empty:
        return flag_new_os, data, data["data_index"], data["Início da Vigência da OS"]

    df = filter_valid_rows(df)

    # converte "Início da Vigência da OS" de dd/mm/aaaa para aaaa-mm-dd
    df["Início da Vigência da OS"] = pd.to_datetime(
        df["Início da Vigência da OS"], format="%d/%m/%Y"
    ).dt.strftime("%Y-%m-%d")

    df["data_index"] = df["Início da Vigência da OS"].astype(str) + "_" + df["index"].astype(str)

    # Ordena por data e index
    df = df.sort_values(by=["data_index"], ascending=True)
    if data_versao_gtfs is not None:
        df = df.loc[(df["Início da Vigência da OS"] == data_versao_gtfs)]

    elif last_captured_os is None:
        last_captured_os = df["data_index"].max()
        df = df.loc[(df["data_index"] == last_captured_os)]

    else:
        # Filtra linhas onde 'data_index' é maior que o último capturado
        df = df.loc[(df["data_index"] > last_captured_os)]

    log(f"Os info: {df.head()}")
    if len(df) >= 1:
        log("Nova OS encontrada!")
        data = df.to_dict(orient="records")[0]  # Converte o DataFrame para um dicionário
        flag_new_os = True  # Se houver mais de uma OS, é uma nova OS

        log(f"OS selecionada: {data}")
    return flag_new_os, data, data["data_index"], data["Início da Vigência da OS"]


@task(nout=2)
def get_raw_gtfs_files(
    os_control,
    local_filepath: list,
    regular_sheet_index: int = None,
    upload_from_gcs: bool = False,
    data_versao_gtfs: str = None,
    dict_gtfs: dict | None = None,
):
    """
    Downloads raw files and processes them.

    Args:
        os_control (dict): A dictionary containing information about the OS (Ordem de Serviço).
        local_filepath (list): A list of local file paths where the downloaded files will be saved.
        regular_sheet_index (int, optional): The index of the regular sheet. Defaults to None.
        upload_from_gcs (bool, optional):
            A boolean indicating whether the files should be uploaded from GCS. Defaults to False.

    Returns:
        raw_filepaths (list): A list of file paths where the downloaded raw files are saved.
        primary_keys (list[list]): A list with the primary_keys for the tables.
    """

    raw_filepaths = []

    log(f"Baixando arquivos: {os_control}")

    if upload_from_gcs:
        log("Baixando arquivos através do GCS")

        # Baixa planilha de OS
        file_bytes_os = io.BytesIO(
            get_upload_storage_blob(
                dataset_id=constants.GTFS_DATASET_ID.value, filename="os"
            ).download_as_bytes()
        )

        # Baixa GTFS
        file_bytes_gtfs = io.BytesIO(
            get_upload_storage_blob(
                dataset_id=constants.GTFS_DATASET_ID.value, filename="gtfs"
            ).download_as_bytes()
        )

    else:
        log("Baixando arquivos através do Google Drive")

        # Criar o serviço da API Google Drive e Google Sheets
        drive_service = get_google_api_service(service_name="drive", version="v3")

        # Baixa planilha de OS
        file_link = os_control["Link da OS"]
        file_bytes_os = download_xlsx(file_link=file_link, drive_service=drive_service)

        # Baixa GTFS
        file_link = os_control["Link do GTFS"]
        file_bytes_gtfs = download_file(file_link=file_link, drive_service=drive_service)

    # Salva os nomes das planilhas
    sheetnames = xl.load_workbook(file_bytes_os).sheetnames
    sheetnames = [name for name in sheetnames if "ANEXO" in name]
    log(f"tabs encontradas na planilha Controle OS: {sheetnames}")

    with zipfile.ZipFile(file_bytes_gtfs, "r") as zipped_file:
        for filename in list(dict_gtfs.keys()):
            if filename == "ordem_servico":
                processa_ordem_servico(
                    sheetnames=sheetnames,
                    file_bytes=file_bytes_os,
                    local_filepath=local_filepath,
                    raw_filepaths=raw_filepaths,
                    regular_sheet_index=regular_sheet_index,
                )
            elif "ordem_servico_trajeto_alternativo" in filename:
                processa_ordem_servico_trajeto_alternativo(
                    sheetnames=sheetnames,
                    file_bytes=file_bytes_os,
                    local_filepath=local_filepath,
                    raw_filepaths=raw_filepaths,
                    data_versao_gtfs=data_versao_gtfs,
                )
            elif "ordem_servico_faixa_horaria" in filename:
                processa_ordem_servico_faixa_horaria(
                    sheetnames=sheetnames,
                    file_bytes=file_bytes_os,
                    local_filepath=local_filepath,
                    raw_filepaths=raw_filepaths,
                    data_versao_gtfs=data_versao_gtfs,
                )

            else:
                # Processa arquivos do GTFS
                data = zipped_file.read(filename + ".txt")
                data = data.decode(encoding="utf-8")

                # encontra a partição correta
                local_file_path = list(filter(lambda x: filename + "/" in x, local_filepath))[0]

                raw_file_path = save_raw_local_func(
                    data=data, filepath=local_file_path, filetype="txt"
                )
                log(f"Saved file: {raw_file_path}")

                raw_filepaths.append(raw_file_path)

    return raw_filepaths, list(dict_gtfs.values())


@task
def upload_raw_data_to_gcs(
    env: str, table_id: str, raw_filepath: str, dataset_id: str, partitions: str
):

    Storage(env=env, dataset_id=dataset_id, table_id=table_id).upload_file(
        mode="raw",
        filepath=raw_filepath,
        partition=partitions,
    )


@task
def upload_staging_data_to_gcs(
    env: str, table_id: str, staging_filepath: str, dataset_id: str, partitions: str
):
    dataset_id_staging = f"{dataset_id}_staging"
    tb_obj = BQTable(env=env, dataset_id=dataset_id_staging, table_id=table_id)

    create_func = partial(
        create_bq_external_table,
        table_obj=tb_obj,
        path=staging_filepath,
        bucket_name=tb_obj.bucket_name,
    )

    append_func = partial(
        Storage(env=env, dataset_id=dataset_id, table_id=table_id).upload_file,
        mode="staging",
        filepath=staging_filepath,
        partition=partitions,
    )

    if not tb_obj.exists():
        log(f"Tabela {tb_obj.table_full_name} não existe. Criando tabela...")
        create_func()
        log(f"Tabela {tb_obj.table_full_name} criada com sucesso.")
    else:
        log(f"Tabela {tb_obj.table_full_name} já existe.")
        append_func()
        log(f"Tabela {tb_obj.table_full_name} atualizada com sucesso.")


@task
def filter_gtfs_table_ids(data_versao_gtfs_str, gtfs_table_capture_params):
    """
    +    Filtra os IDs de tabelas GTFS com base na versão dos dados.
    +
    +    Args:
    +        data_versao_gtfs_str (str): Data de versão do GTFS no formato 'YYYY-MM-DD'.
    +        gtfs_table_capture_params (dict): Dicionário com os parâmetros de captura das tabelas.
    +
    +    Returns:
    +        dict: Dicionário filtrado com os parâmetros de captura das tabelas.
    +"""
    if data_versao_gtfs_str >= constants.DATA_GTFS_V2_INICIO.value:
        gtfs_table_capture_params.pop("ordem_servico", None)

    if data_versao_gtfs_str < constants.DATA_GTFS_V4_INICIO.value:
        gtfs_table_capture_params.pop("ordem_servico_faixa_horaria_sentido", None)

    if data_versao_gtfs_str >= constants.DATA_GTFS_V4_INICIO.value:
        gtfs_table_capture_params.pop("ordem_servico_faixa_horaria", None)

    if data_versao_gtfs_str < constants.DATA_GTFS_V5_INICIO.value:
        gtfs_table_capture_params.pop("ordem_servico_trajeto_alternativo_sentido", None)

    if data_versao_gtfs_str >= constants.DATA_GTFS_V5_INICIO.value:
        gtfs_table_capture_params.pop("ordem_servico_trajeto_alternativo", None)

    return gtfs_table_capture_params
