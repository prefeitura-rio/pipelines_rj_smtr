# -*- coding: utf-8 -*-
"""
Utils for gtfs
"""
import io
import re

import openpyxl as xl
import pandas as pd
import requests
from googleapiclient.http import MediaIoBaseDownload
from prefeitura_rio.pipelines_utils.logging import log
from unidecode import unidecode

from pipelines.constants import constants
from pipelines.migration.utils import save_raw_local_func


def filter_valid_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filter the valid rows in a DataFrame based on specific conditions.

    Args:
        df (pd.DataFrame): The input DataFrame to filter.

    Returns:
        pd.DataFrame: The filtered DataFrame.

    """
    # salva index atual em uma coluna
    df["index"] = df.index
    # Remove linhas com valores nulos
    df.dropna(how="all", inplace=True)

    # Remove linhas onde 'Fim da Vigência da OS' == 'Sem Vigência'
    df = df[df["Fim da Vigência da OS"] != "Sem Vigência"]

    # Remove linhas onde 'Submeter mudanças para Dados' == False
    df = df[df["Submeter mudanças para Dados"] == True]  # noqa

    # Remove linhas onde 'Arquivo OS' e 'Arquivo GTFS' são nulos
    df = df[~df["Início da Vigência da OS"].isnull()]
    df = df[~df["Arquivo OS"].isnull()]
    df = df[~df["Arquivo GTFS"].isnull()]
    df = df[~df["Link da OS"].isnull()]
    df = df[~df["Link do GTFS"].isnull()]
    return df


def download_controle_os_csv(url):
    """
    Downloads a CSV file from the given URL and returns its content as a pandas DataFrame.

    Args:
        url (str): The URL of the CSV file to download.

    Returns:
        pandas.DataFrame: The content of the downloaded CSV file as a DataFrame.

    """

    response = requests.get(url=url, timeout=constants.MAX_TIMEOUT_SECONDS.value)
    response.raise_for_status()  # Verifica se houve algum erro na requisição
    response.encoding = "utf-8"
    # Carrega o conteúdo da resposta em um DataFrame
    df = pd.read_csv(io.StringIO(response.text))

    log(f"Download concluído! Dados:\n{df.head()}")
    return df


def convert_to_float(value):
    """
    Converts a string value to a float.

    Args:
        value (str): The string value to be converted.

    Returns:
        float: The converted float value.
    """
    if "," in value:
        value = value.replace(".", "").replace(",", ".").strip()
    return float(value)


def processa_OS(file_link: str, drive_service, local_filepath: list, raw_filepaths: list):
    """
    Process the OS data from an Excel file.

    Args:
        file_link (str): The link to the Excel file.
        drive_service: The drive service object for downloading the file.
        local_filepath (list): The list of local file paths.
        raw_filepaths (list): The list of raw file paths.

    Returns:
        None
    """

    file_bytes = download_xlsx(file_link=file_link, drive_service=drive_service)

    # Processar a planilha
    wb = xl.load_workbook(file_bytes)
    sheetnames = wb.sheetnames

    processa_ordem_servico(
        sheetnames=sheetnames,
        file_bytes=file_bytes,
        local_filepath=local_filepath,
        raw_filepaths=raw_filepaths,
    )

    processa_ordem_servico_trajeto_alternativo(
        sheetnames=sheetnames,
        file_bytes=file_bytes,
        local_filepath=local_filepath,
        raw_filepaths=raw_filepaths,
    )


def download_xlsx(file_link, drive_service):
    """
    Downloads an XLSX file from Google Drive.

    Args:
        file_link (str): The link to the file in Google Drive.
        drive_service: The Google Drive service object.

    Returns:
        io.BytesIO: The downloaded XLSX file as a BytesIO object.
    """
    file_id = file_link.split("/")[-2]

    file = (
        drive_service.files().get(fileId=file_id, supportsAllDrives=True).execute()
    )  # pylint: disable=E1101
    mime_type = file.get("mimeType")

    if "google-apps" in mime_type:
        request = drive_service.files().export(  # pylint: disable=E1101
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = drive_service.files().get_media(fileId=file_id)  # pylint: disable=E1101

    file_bytes = io.BytesIO()
    downloader = MediaIoBaseDownload(file_bytes, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()

    file_bytes.seek(0)
    return file_bytes


def normalizar_horario(horario):
    """
    Normalizes the given time string.

    If the time string contains the word "day", it splits the string into days and time.
    It converts the days, hours, minutes, and seconds into total hours and returns
    the normalized time string.

    If the time string does not contain the word "day", it returns the time string as is.

    Args:
        horario (str): The time string to be normalized.

    Returns:
        str: The normalized time string.

    """
    if "day" in horario:
        days, time = horario.split(", ")
        days = int(days.split(" ")[0])
        hours, minutes, seconds = map(int, time.split(":"))
        total_hours = days * 24 + hours
        return f"{total_hours:02}:{minutes:02}:{seconds:02}"
    else:
        return horario.split(" ")[1] if " " in horario else horario


def processa_ordem_servico(
    sheetnames, file_bytes, local_filepath, raw_filepaths, regular_sheet_index=None
):
    """
    Process the OS data from an Excel file.

    Args:
        sheetnames (list): List of sheet names in the Excel file.
        file_bytes (bytes): The Excel file in bytes format.
        local_filepath (str): The local file path where the processed data will be saved.
        raw_filepaths (list): List of raw file paths.
        regular_sheet_index (int, optional): The index of the regular sheet. Defaults to 0.

    Raises:
        Exception: If there are more than 2 tabs in the file or if there are missing or
        duplicated columns in the order of service data.
        Exception: If the validation of 'km_test' and 'km_dia_util' fails.

    Returns:
        None
    """

    sheets = [(i, name) for i, name in enumerate(sheetnames) if "ANEXO I " in name]
    if not sheets:
        raise ValueError("Nenhuma aba 'ANEXO I' encontrada no arquivo.")
    sheets_data = []

    columns = {
        "Serviço": "servico",
        "Vista": "vista",
        "Consórcio": "consorcio",
        "Extensão de Ida": "extensao_ida",
        "Extensão de Volta": "extensao_volta",
        "Horário Inicial": "horario_inicio",
        "Horário\nInicial": "horario_inicio",
        "Horário Fim": "horario_fim",
        "Horário\nFim": "horario_fim",
        "Partidas Ida Dia Útil": "partidas_ida_du",
        "Partidas Volta Dia Útil": "partidas_volta_du",
        "Viagens Dia Útil": "viagens_du",
        "Quilometragem Dia Útil": "km_dia_util",
        "KM Dia Útil": "km_dia_util",
        "Partidas Ida Sábado": "partidas_ida_sabado",
        "Partidas Volta Sábado": "partidas_volta_sabado",
        "Viagens Sábado": "viagens_sabado",
        "Quilometragem Sábado": "km_sabado",
        "KM Sábado": "km_sabado",
        "Partidas Ida Domingo": "partidas_ida_domingo",
        "Partidas Volta Domingo": "partidas_volta_domingo",
        "Viagens Domingo": "viagens_domingo",
        "Quilometragem Domingo": "km_domingo",
        "KM Domingo": "km_domingo",
        "Partidas Ida Ponto Facultativo": "partidas_ida_pf",
        "Partidas Volta Ponto Facultativo": "partidas_volta_pf",
        "Viagens Ponto Facultativo": "viagens_pf",
        "Quilometragem Ponto Facultativo": "km_pf",
        "KM Ponto Facultativo": "km_pf",
        "tipo_os": "tipo_os",
    }

    for sheet_index, sheet_name in sheets:
        log(f"########## {sheet_name} ##########")

        match = re.search(r"\((.*?)\)", sheet_name)
        if not match:
            raise ValueError(f"Não foi possível extrair tipo_os do nome da aba: {sheet_name}")
        tipo_os = match.group(1)

        quadro = pd.read_excel(file_bytes, sheet_name=sheet_name, dtype=object)

        quadro = quadro.rename(columns=columns)

        quadro["servico"] = quadro["servico"].astype(str)
        quadro["servico"] = quadro["servico"].str.extract(r"([A-Z]+)", expand=False).fillna(
            ""
        ) + quadro["servico"].str.extract(r"([0-9]+)", expand=False).fillna("")

        quadro["tipo_os"] = tipo_os

        quadro = quadro[list(set(columns.values()))]
        quadro = quadro.replace("—", 0)
        quadro = quadro.reindex(columns=list(set(columns.values())))

        hora_cols = [coluna for coluna in quadro.columns if "horario" in coluna]
        quadro[hora_cols] = quadro[hora_cols].astype(str)

        for hora_col in hora_cols:
            quadro[hora_col] = quadro[hora_col].apply(normalizar_horario)

        cols = [
            coluna
            for coluna in quadro.columns
            if "km" in coluna or "viagens" in coluna or "partida" in coluna
        ]

        for col in cols:
            quadro[col] = quadro[col].astype(str).apply(convert_to_float).astype(float).fillna(0)

        extensao_cols = ["extensao_ida", "extensao_volta"]
        quadro[extensao_cols] = quadro[extensao_cols].astype(str)
        for col in extensao_cols:
            quadro[col] = quadro[col].str.replace(".", "", regex=False)
        quadro[extensao_cols] = quadro[extensao_cols].apply(pd.to_numeric)

        quadro["extensao_ida"] = quadro["extensao_ida"] / 1000
        quadro["extensao_volta"] = quadro["extensao_volta"] / 1000

        sheets_data.append(quadro)

    quadro_geral = pd.concat(sheets_data, ignore_index=True)

    columns_in_dataframe = set(quadro_geral.columns)
    columns_in_values = set(list(columns.values()))

    all_columns_present = columns_in_dataframe.issubset(columns_in_values)
    no_duplicate_columns = len(columns_in_dataframe) == len(quadro_geral.columns)
    missing_columns = columns_in_values - columns_in_dataframe

    log(
        f"All columns present: {all_columns_present}/"
        f"No duplicate columns: {no_duplicate_columns}/"
        f"Missing columns: {missing_columns}"
    )

    if not all_columns_present or not no_duplicate_columns:
        raise Exception("Missing or duplicated columns in ordem_servico")

    local_file_path = list(filter(lambda x: "ordem_servico" in x, local_filepath))[0]
    quadro_geral_csv = quadro_geral.to_csv(index=False)
    raw_file_path = save_raw_local_func(
        data=quadro_geral_csv, filepath=local_file_path, filetype="csv"
    )
    log(f"Saved file: {raw_file_path}")

    raw_filepaths.append(raw_file_path)


def processa_ordem_servico_trajeto_alternativo(
    sheetnames,
    file_bytes,
    local_filepath,
    raw_filepaths,
    data_versao_gtfs,
):
    """
    Process 'Trajetos Alternativos' from an Excel file.

    Args:
        sheetnames (list): List of sheet names in the Excel file.
        file_bytes (bytes): Bytes of the Excel file.
        local_filepath (str): Local file path.
        raw_filepaths (list): List of raw file paths.

    Returns:
        None

    Raises:
        Exception: If there are missing or duplicated columns in 'Trajetos Alternativos'.
    """

    sheets = [(i, name) for i, name in enumerate(sheetnames) if "ANEXO II " in name]
    if not sheets:
        raise ValueError("Nenhuma aba 'ANEXO II' encontrada no arquivo.")
    sheets_data = []

    alt_columns = {}

    if data_versao_gtfs < constants.DATA_GTFS_V5_INICIO.value:
        alt_columns = {
            "Serviço": "servico",
            "Vista": "vista",
            "Consórcio": "consorcio",
            "Extensão de Ida": "extensao_ida",
            "Extensão\nde Ida": "extensao_ida",
            "Extensão de Volta": "extensao_volta",
            "Extensão\nde Volta": "extensao_volta",
            "Evento": "evento",
            "Horário Inicial Interdição": "inicio_periodo",
            "Horário Final Interdição": "fim_periodo",
            "Descrição": "descricao",
            "Ativação": "ativacao",
            "tipo_os": "tipo_os",
        }
    else:
        alt_columns = {
            "Serviço": "servico",
            "Vista": "vista",
            "Sentido": "sentido",
            "Extensão": "extensao",
            "Consórcio": "consorcio",
            "Evento": "evento",
            "Descrição": "descricao",
            "Ativação": "ativacao",
            "tipo_os": "tipo_os",
        }

    for sheet_index, sheet_name in sheets:
        log(f"########## {sheet_name} ##########")
        match = re.search(r"\((.*?)\)", sheet_name)
        if not match:
            raise ValueError(f"Não foi possível extrair tipo_os do nome da aba: {sheet_name}")
        tipo_os = match.group(1)

        df = pd.read_excel(file_bytes, sheet_name=sheet_name, dtype=object)

        df = df.rename(columns=alt_columns)
        df["tipo_os"] = tipo_os
        sheets_data.append(df)

    ordem_servico_trajeto_alternativo = pd.concat(sheets_data, ignore_index=True)

    columns_in_dataframe = set(ordem_servico_trajeto_alternativo.columns)
    columns_in_values = set(list(alt_columns.values()))

    all_columns_present = columns_in_dataframe.issubset(columns_in_values)
    no_duplicate_columns = len(columns_in_dataframe) == len(
        ordem_servico_trajeto_alternativo.columns
    )
    missing_columns = columns_in_values - columns_in_dataframe

    log(
        f"All columns present: {all_columns_present}/"
        f"No duplicate columns: {no_duplicate_columns}/"
        f"Missing columns: {missing_columns}"
    )

    if not all_columns_present or not no_duplicate_columns:
        raise Exception("Missing or duplicated columns in ordem_servico_trajeto_alternativo")

    local_file_path = list(
        filter(lambda x: "ordem_servico_trajeto_alternativo" in x, local_filepath)
    )[0]
    ordem_servico_trajeto_alternativo_csv = ordem_servico_trajeto_alternativo.to_csv(index=False)
    raw_file_path = save_raw_local_func(
        data=ordem_servico_trajeto_alternativo_csv,
        filepath=local_file_path,
        filetype="csv",
    )
    log(f"Saved file: {raw_file_path}")

    raw_filepaths.append(raw_file_path)


def download_file(file_link, drive_service):
    """
    Downloads a file from Google Drive.

    Args:
        file_link (str): The link to the file in Google Drive.
        drive_service: The Google Drive service object.

    Returns:
        io.BytesIO: The downloaded file as a BytesIO object.
    """
    file_id = file_link.split("/")[-2]

    request = drive_service.files().get_media(fileId=file_id, supportsAllDrives=True)
    file_bytes = io.BytesIO()
    downloader = MediaIoBaseDownload(file_bytes, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    return file_bytes


def processa_ordem_servico_faixa_horaria(
    sheetnames, file_bytes, local_filepath, raw_filepaths, data_versao_gtfs
):
    if data_versao_gtfs >= constants.DATA_GTFS_V2_INICIO.value:
        sheets = [(i, name) for i, name in enumerate(sheetnames) if "ANEXO I " in name]
        if not sheets:
            raise ValueError("Nenhuma aba 'ANEXO I' encontrada no arquivo.")
    else:
        sheets = [(i, name) for i, name in enumerate(sheetnames) if "ANEXO III " in name]
        if not sheets:
            raise ValueError("Nenhuma aba 'ANEXO III' encontrada no arquivo.")
    sheets_data = []

    columns = {
        "Serviço": "servico",
        "Vista": "vista",
        "Consórcio": "consorcio",
        "Extensão de Ida": "extensao_ida",
        "Extensão de Volta": "extensao_volta",
        "Horário Inicial Dias Úteis": "horario_inicio_dias_uteis",
        "Horário Fim Dias Úteis": "horario_fim_dias_uteis",
        "Horário Inicial - Dias Úteis": "horario_inicio_dias_uteis",
        "Horário Fim - Dias Úteis": "horario_fim_dias_uteis",
        "Partidas Ida - Dias Úteis": "partidas_ida_dias_uteis",
        "Partidas Volta - Dias Úteis": "partidas_volta_dias_uteis",
        "Viagens - Dias Úteis": "viagens_dias_uteis",
        "Quilometragem - Dias Úteis": "quilometragem_dias_uteis",
        "KM - Dias Úteis": "quilometragem_dias_uteis",
        "Horário Inicial Sábado": "horario_inicio_sabado",
        "Horário Fim Sábado": "horario_fim_sabado",
        "Horário Inicial - Sábado": "horario_inicio_sabado",
        "Horário Fim - Sábado": "horario_fim_sabado",
        "Partidas Ida - Sábado": "partidas_ida_sabado",
        "Partidas Volta - Sábado": "partidas_volta_sabado",
        "Viagens - Sábado": "viagens_sabado",
        "Quilometragem - Sábado": "quilometragem_sabado",
        "KM - Sábado": "quilometragem_sabado",
        "Horário Inicial Domingo": "horario_inicio_domingo",
        "Horário Fim Domingo": "horario_fim_domingo",
        "Horário Inicial - Domingo": "horario_inicio_domingo",
        "Horário Fim - Domingo": "horario_fim_domingo",
        "Partidas Ida - Domingo": "partidas_ida_domingo",
        "Partidas Volta - Domingo": "partidas_volta_domingo",
        "Viagens - Domingo": "viagens_domingo",
        "Quilometragem - Domingo": "quilometragem_domingo",
        "KM - Domingo": "quilometragem_domingo",
        "Horário Inicial Ponto Facultativo": "horario_inicio_ponto_facultativo",
        "Horário Fim Ponto Facultativo": "horario_fim_ponto_facultativo",
        "Horário Inicial - Ponto Facultativo": "horario_inicio_ponto_facultativo",
        "Horário Fim - Ponto Facultativo": "horario_fim_ponto_facultativo",
        "Partidas Ida - Ponto Facultativo": "partidas_ida_ponto_facultativo",
        "Partidas Volta - Ponto Facultativo": "partidas_volta_ponto_facultativo",
        "Viagens - Ponto Facultativo": "viagens_ponto_facultativo",
        "Quilometragem - Ponto Facultativo": "quilometragem_ponto_facultativo",
        "KM - Ponto Facultativo": "quilometragem_ponto_facultativo",
        "tipo_os": "tipo_os",
    }

    metricas = ["Partidas", "Partidas Ida", "Partidas Volta", "Quilometragem", "KM"]
    dias = ["Dias Úteis", "Sábado", "Domingo", "Ponto Facultativo"]
    formatos = [
        "{metrica} entre {intervalo} — {dia}",
        "{metrica} entre {intervalo} - {dia}",
        "{metrica} entre {intervalo} ({dia})",
        "{metrica} {intervalo} - {dia}",
    ]

    if data_versao_gtfs < constants.DATA_GTFS_V3_INICIO.value:
        intervalos = [
            "00h e 03h",
            "03h e 12h",
            "12h e 21h",
            "21h e 24h",
            "24h e 03h (dia seguinte)",
        ]
    elif (
        data_versao_gtfs >= constants.DATA_GTFS_V3_INICIO.value
        and data_versao_gtfs < constants.DATA_GTFS_V4_INICIO.value
    ):
        intervalos = [
            "00h e 03h",
            "03h e 06h",
            "06h e 09h",
            "09h e 12h",
            "12h e 15h",
            "15h e 18h",
            "18h e 21h",
            "21h e 24h",
            "24h e 03h (dia seguinte)",
        ]
    else:
        intervalos = [
            "00h à 01h",
            "01h à 02h",
            "02h à 03h",
            "03h à 04h",
            "04h à 05h",
            "05h à 06h",
            "06h à 09h",
            "09h à 12h",
            "12h à 15h",
            "15h à 18h",
            "18h à 21h",
            "21h à 22h",
            "22h à 23h",
            "23h à 24h",
        ]

    fh_columns = {
        formato.format(metrica=metrica, intervalo=intervalo, dia=dia): unidecode(
            (
                "quilometragem"
                if metrica in ["Quilometragem", "KM"]
                else (
                    "partidas"
                    if metrica == "Partidas"
                    and data_versao_gtfs < constants.DATA_GTFS_V2_INICIO.value
                    or data_versao_gtfs >= constants.DATA_GTFS_V4_INICIO.value
                    else ("partidas_ida" if metrica == "Partidas Ida" else "partidas_volta")
                )
            )
            + f"_entre_{intervalo.replace(' ', '_').replace('(', '').replace(')', '').replace('-', '_')}_{dia.lower().replace(' ', '_')}"  # noqa
        )
        for metrica in metricas
        for intervalo in intervalos
        for dia in dias
        for formato in formatos
    }

    if data_versao_gtfs >= constants.DATA_GTFS_V4_INICIO.value:
        fh_columns["Serviço"] = "servico"
        fh_columns["Consórcio"] = "consorcio"
        fh_columns["tipo_os"] = "tipo_os"
        fh_columns["Sentido"] = "sentido"
        fh_columns["Extensão"] = "extensao"
        fh_columns["Vista"] = "vista"
        columns = fh_columns.copy()
    elif data_versao_gtfs >= constants.DATA_GTFS_V2_INICIO.value:
        columns.update(fh_columns)
    else:
        fh_columns["Serviço"] = "servico"
        fh_columns["Consórcio"] = "consorcio"
        fh_columns["tipo_os"] = "tipo_os"
        columns = fh_columns.copy()

    for sheet_index, sheet_name in sheets:
        log(f"########## {sheet_name} ##########")
        match = re.search(r"\((.*?)\)", sheet_name)
        if not match:
            raise ValueError(f"Não foi possível extrair tipo_os do nome da aba: {sheet_name}")
        tipo_os = match.group(1)

        df = pd.read_excel(file_bytes, sheet_name=sheet_name, dtype=object)

        df.columns = (
            df.columns.str.replace("\n", " ").str.strip().str.replace(r"\s+", " ", regex=True)
        )
        df = df.rename(columns=lambda x: x.replace("Dia Útil", "Dias Úteis"))
        df = df.rename(columns=columns)

        for col in df.columns:
            if "quilometragem" in col or "viagens" in col or "partidas" in col:
                df[col] = df[col].astype(str).replace(["—", "-"], 0)
            if "quilometragem" in col or "viagens" in col:
                df[col] = df[col].astype(str).apply(convert_to_float).astype(float)
            if "extensao" in col:
                df[col] = df[col].apply(pd.to_numeric)
                df[col] = df[col] / 1000
            if "horario" in col:
                df[col] = df[col].astype(str)
                df[col] = df[col].apply(normalizar_horario)

        df["tipo_os"] = tipo_os

        aux_columns_in_dataframe = set(df.columns)
        columns_in_values = set(list(columns.values()))
        aux_missing_columns = columns_in_values - aux_columns_in_dataframe

        for coluna in aux_missing_columns:
            if "ponto_facultativo" in coluna:
                df[coluna] = 0

        sheets_data.append(df)

    ordem_servico_faixa_horaria = pd.concat(sheets_data, ignore_index=True)

    columns_in_dataframe = set(ordem_servico_faixa_horaria.columns)
    missing_columns = columns_in_values - columns_in_dataframe
    all_columns_present = columns_in_dataframe.issubset(columns_in_values)
    no_duplicate_columns = len(columns_in_dataframe) == len(ordem_servico_faixa_horaria.columns)

    log(
        f"All columns present: {all_columns_present}\n"
        f"No duplicate columns: {no_duplicate_columns}\n"
        f"Missing columns: {missing_columns}"
    )

    if not all_columns_present or not no_duplicate_columns:
        log(columns_in_values.difference(columns_in_dataframe))
        log(columns_in_dataframe.difference(columns_in_values))
        raise Exception("Missing or duplicated columns in ordem_servico_faixa_horaria")

    local_file_path = list(filter(lambda x: "ordem_servico_faixa_horaria" in x, local_filepath))[0]
    ordem_servico_faixa_horaria_csv = ordem_servico_faixa_horaria.to_csv(index=False)
    raw_file_path = save_raw_local_func(
        data=ordem_servico_faixa_horaria_csv,
        filepath=local_file_path,
        filetype="csv",
    )
    log(f"Saved file: {raw_file_path}")

    raw_filepaths.append(raw_file_path)
