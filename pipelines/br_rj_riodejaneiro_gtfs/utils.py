# -*- coding: utf-8 -*-
import io

import openpyxl as xl
import pandas as pd
import requests
from googleapiclient.http import MediaIoBaseDownload
from prefeitura_rio.pipelines_utils.logging import log

from pipelines.constants import constants
from pipelines.utils.backup.utils import save_raw_local_func


def filter_valid_rows(df: pd.DataFrame) -> pd.DataFrame:

    # salva index atual em uma coluna
    df["index"] = df.index
    # Remove linhas com valores nulos
    df.dropna(how="all", inplace=True)

    # Remove linhas onde 'Fim da Vigência da OS' == 'Sem Vigência'
    df = df[df["Fim da Vigência da OS"] != "Sem Vigência"]

    # Remove linhas onde 'Submeter mudanças para Dados' == False
    df = df[df["Submeter mudanças para Dados"] == True]

    # Remove linhas onde 'Arquivo OS' e 'Arquivo GTFS' são nulos
    df = df[~df["Início da Vigência da OS"].isnull()]
    df = df[~df["Arquivo OS"].isnull()]
    df = df[~df["Arquivo GTFS"].isnull()]
    df = df[~df["Link da OS"].isnull()]
    df = df[~df["Link do GTFS"].isnull()]
    return df


def download_controle_os_csv(url):

    response = requests.get(url=url, timeout=constants.MAX_TIMEOUT_SECONDS.value)
    response.raise_for_status()  # Verifica se houve algum erro na requisição
    response.encoding = "utf-8"
    # Carrega o conteúdo da resposta em um DataFrame
    df = pd.read_csv(io.StringIO(response.text))

    log(f"Download concluído! Dados:\n{df.head()}")
    return df


def convert_to_float(value):
    if "," in value:
        value = value.replace(".", "").replace(",", ".").strip()
    return float(value)


def processa_OS(file_link: str, drive_service, local_filepath: list, raw_filepaths: list):

    file_bytes = download_xlsx(file_link=file_link, drive_service=drive_service)

    # Processar a planilha
    wb = xl.load_workbook(file_bytes)
    sheetnames = wb.sheetnames

    processa_ordem_servico(
        sheetnames=sheetnames,
        file_bytes=file_bytes,
        local_filepath=local_filepath,
        raw_filepaths=raw_filepaths,
    )

    processa_ordem_servico_trajeto_alternativo(
        sheetnames=sheetnames,
        file_bytes=file_bytes,
        local_filepath=local_filepath,
        raw_filepaths=raw_filepaths,
    )


def download_xlsx(file_link, drive_service):
    file_id = file_link.split("/")[-2]

    file = drive_service.files().get(fileId=file_id).execute()  # pylint: disable=E1101
    mime_type = file.get("mimeType")

    if "google-apps" in mime_type:
        request = drive_service.files().export(  # pylint: disable=E1101
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = drive_service.files().get_media(fileId=file_id)  # pylint: disable=E1101

    file_bytes = io.BytesIO()
    downloader = MediaIoBaseDownload(file_bytes, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()

    file_bytes.seek(0)
    return file_bytes


def processa_ordem_servico(sheetnames, file_bytes, local_filepath, raw_filepaths):
    # conta quantos sheets tem no arquivo sem o nome de Anexo II
    sheets_range = len(sheetnames) - len([x for x in sheetnames if "ANEXO II" in x])
    quadro_geral = pd.DataFrame()

    for i in range(0, sheets_range):
        log(f"########## {sheetnames[i]} ##########")
        quadro = pd.read_excel(file_bytes, sheet_name=sheetnames[i], dtype=object)

        columns = {
            "Serviço": "servico",
            "Vista": "vista",
            "Consórcio": "consorcio",
            "Extensão de Ida": "extensao_ida",
            "Extensão de Volta": "extensao_volta",
            "Horário Inicial": "horario_inicio",
            "Horário\nInicial": "horario_inicio",
            "Horário Fim": "horario_fim",
            "Horário\nFim": "horario_fim",
            "Partidas Ida Dia Útil": "partidas_ida_du",
            "Partidas Volta Dia Útil": "partidas_volta_du",
            "Viagens Dia Útil": "viagens_du",
            "Quilometragem Dia Útil": "km_dia_util",
            "Partidas Ida Sábado": "partidas_ida_sabado",
            "Partidas Volta Sábado": "partidas_volta_sabado",
            "Viagens Sábado": "viagens_sabado",
            "Quilometragem Sábado": "km_sabado",
            "Partidas Ida Domingo": "partidas_ida_domingo",
            "Partidas Volta Domingo": "partidas_volta_domingo",
            "Viagens Domingo": "viagens_domingo",
            "Quilometragem Domingo": "km_domingo",
            "Partidas Ida Ponto Facultativo": "partidas_ida_pf",
            "Partidas Volta Ponto Facultativo": "partidas_volta_pf",
            "Viagens Ponto Facultativo": "viagens_pf",
            "Quilometragem Ponto Facultativo": "km_pf",
        }

        quadro = quadro.rename(columns=columns)

        quadro["servico"] = quadro["servico"].astype(str)
        quadro["servico"] = quadro["servico"].str.extract(r"([A-Z]+)", expand=False).fillna(
            ""
        ) + quadro["servico"].str.extract(r"([0-9]+)", expand=False).fillna("")

        quadro = quadro[list(set(columns.values()))]
        quadro = quadro.replace("—", 0)
        quadro = quadro.reindex(columns=list(set(columns.values())))

        hora_cols = [coluna for coluna in quadro.columns if "horario" in coluna]
        quadro[hora_cols] = quadro[hora_cols].astype(str)

        for hora_col in hora_cols:
            quadro[hora_col] = quadro[hora_col].apply(lambda x: x.split(" ")[1] if " " in x else x)

        cols = [
            coluna
            for coluna in quadro.columns
            if "km" in coluna or "viagens" in coluna or "partida" in coluna
        ]

        for col in cols:
            quadro[col] = quadro[col].astype(str).apply(convert_to_float).astype(float).fillna(0)

        extensao_cols = ["extensao_ida", "extensao_volta"]
        quadro[extensao_cols] = quadro[extensao_cols].astype(str)
        quadro[extensao_cols] = quadro[extensao_cols].apply(pd.to_numeric)

        quadro["extensao_ida"] = quadro["extensao_ida"] / 1000
        quadro["extensao_volta"] = quadro["extensao_volta"] / 1000

        if i == 0:
            quadro["tipo_os"] = "Regular"

        quadro_geral = pd.concat([quadro_geral, quadro])

    # Verificações
    columns_in_dataframe = set(quadro_geral.columns)
    columns_in_values = set(list(columns.values()) + ["tipo_os"])

    all_columns_present = columns_in_dataframe.issubset(columns_in_values)
    no_duplicate_columns = len(columns_in_dataframe) == len(quadro_geral.columns)
    missing_columns = columns_in_values - columns_in_dataframe

    log(
        f"All columns present: {all_columns_present}/"
        f"No duplicate columns: {no_duplicate_columns}/"
        f"Missing columns: {missing_columns}"
    )

    if not all_columns_present or not no_duplicate_columns:
        raise Exception("Missing or duplicated columns in ordem_servico")

    quadro_test = quadro_geral.copy()
    quadro_test["km_test"] = round(
        (quadro_geral["partidas_volta_du"] * quadro_geral["extensao_volta"])
        + (quadro_geral["partidas_ida_du"] * quadro_geral["extensao_ida"]),
        2,
    )
    quadro_test["dif"] = quadro_test["km_test"] - quadro_test["km_dia_util"]

    if not (
        round(abs(quadro_test["dif"].max()), 2) <= 0.01
        and round(abs(quadro_test["dif"].min()), 2) <= 0.01
    ):
        raise Exception("failed to validate km_test and km_dia_util")

    local_file_path = list(filter(lambda x: "ordem_servico" in x, local_filepath))[0]
    quadro_geral_csv = quadro_geral.to_csv(index=False)
    raw_file_path = save_raw_local_func(
        data=quadro_geral_csv, filepath=local_file_path, filetype="csv"
    )
    log(f"Saved file: {raw_file_path}")

    raw_filepaths.append(raw_file_path)


def processa_ordem_servico_trajeto_alternativo(
    sheetnames, file_bytes, local_filepath, raw_filepaths
):
    # Pre-tratamento para "Trajeto Alternativo"
    sheet = -1
    log(f"########## {sheetnames[sheet]} ##########")

    ordem_servico_trajeto_alternativo = pd.read_excel(
        file_bytes, sheet_name=sheetnames[sheet], dtype=object
    )

    alt_columns = {
        "Serviço": "servico",
        "Vista": "vista",
        "Consórcio": "consorcio",
        "Extensão de Ida": "extensao_ida",
        "Extensão\nde Ida": "extensao_ida",
        "Extensão de Volta": "extensao_volta",
        "Extensão\nde Volta": "extensao_volta",
        "Evento": "evento",
        "Horário Inicial Interdição": "inicio_periodo",
        "Horário Final Interdição": "fim_periodo",
        "Descrição": "descricao",
        "Ativação": "ativacao",
    }

    ordem_servico_trajeto_alternativo = ordem_servico_trajeto_alternativo.rename(
        columns=alt_columns
    )

    columns_in_dataframe = set(ordem_servico_trajeto_alternativo.columns)
    columns_in_values = set(list(alt_columns.values()))

    all_columns_present = columns_in_dataframe.issubset(columns_in_values)
    no_duplicate_columns = len(columns_in_dataframe) == len(
        ordem_servico_trajeto_alternativo.columns
    )
    missing_columns = columns_in_values - columns_in_dataframe

    log(
        f"All columns present: {all_columns_present}/"
        f"No duplicate columns: {no_duplicate_columns}/"
        f"Missing columns: {missing_columns}"
    )

    if not all_columns_present or not no_duplicate_columns:
        raise Exception("Missing or duplicated columns in ordem_servico_trajeto_alternativo")

    local_file_path = list(
        filter(lambda x: "ordem_servico_trajeto_alternativo" in x, local_filepath)
    )[0]
    ordem_servico_trajeto_alternativo_csv = ordem_servico_trajeto_alternativo.to_csv(index=False)
    raw_file_path = save_raw_local_func(
        data=ordem_servico_trajeto_alternativo_csv,
        filepath=local_file_path,
        filetype="csv",
    )
    log(f"Saved file: {raw_file_path}")

    raw_filepaths.append(raw_file_path)
