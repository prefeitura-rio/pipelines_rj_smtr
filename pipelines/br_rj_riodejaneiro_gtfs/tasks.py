# -*- coding: utf-8 -*-
import io
import os
import zipfile
from datetime import datetime

import openpyxl as xl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from prefect import task
from prefeitura_rio.pipelines_utils.logging import log
from prefeitura_rio.pipelines_utils.redis_pal import get_redis_client

from pipelines.br_rj_riodejaneiro_gtfs.utils import (
    download_controle_os_csv,
    download_file,
    download_xlsx,
    filter_valid_rows,
    processa_ordem_servico,
    processa_ordem_servico_trajeto_alternativo,
)
from pipelines.constants import constants
from pipelines.utils.backup.utils import save_raw_local_func


@task
def get_last_capture_os(dataset_id: str, mode: str = "prod") -> dict:
    """
    Retrieves the last captured OS for a given dataset ID and mode.

    Args:
        dataset_id (str): The ID of the dataset.
        mode (str, optional): The mode of operation. Defaults to "prod".

    Returns:
        dict: The last captured OS.

    """
    redis_client = get_redis_client()
    fetch_key = f"{dataset_id}.last_captured_os"
    if mode != "prod":
        fetch_key = f"{mode}.{fetch_key}"

    last_captured_os = redis_client.get(fetch_key)

    log(f"Last captured os: {last_captured_os}")

    return last_captured_os


@task
def update_last_captured_os(dataset_id: str, data_index: str, mode: str = "prod") -> None:
    """
    Update the last captured operating system for a given dataset.

    Args:
        dataset_id (str): The ID of the dataset.
        data_index (str): The last captured operating system.
        mode (str, optional): The mode of operation. Defaults to "prod".

    Returns:
        None
    """
    redis_client = get_redis_client()
    fetch_key = f"{dataset_id}.last_captured_os"
    if mode != "prod":
        fetch_key = f"{mode}.{fetch_key}"

    redis_client.set(fetch_key, {"last_captured_os": data_index})


@task(nout=4)
def get_os_info(last_captured_os: str) -> dict:
    """
    Retrieves information about the OS.

    Args:
        last_captured_os (str): The last captured OS data_index.

    Returns:
        tuple: A tuple containing the following elements:
            - flag_new_os (bool): Indicates whether a new OS was found.
            - data (dict): A dictionary containing the OS information.
            - data_index (str): The index of the captured OS.
            - inicio_vigencia_os (str): The start date of the captured OS.

    """
    df = download_controle_os_csv(constants.GTFS_CONTROLE_OS_URL.value)

    flag_new_os = False
    data = {"Início da Vigência da OS": None, "data_index": None}

    if df.empty:
        return flag_new_os, data, data["data_index"], data["Início da Vigência da OS"]

    df = filter_valid_rows(df)

    df["data_index"] = df["Início da Vigência da OS"].astype(str) + "_" + df["index"].astype(str)

    # Ordena por despacho
    df = df.sort_values(by=["data_index"], ascending=True)
    if last_captured_os is None:
        last_captured_os = df["data_index"].max()
        df = df.loc[(df["data_index"] == last_captured_os)]
    else:
        # Filtra linhas onde 'Despacho' é maior que o último capturado
        df = df.loc[(df["data_index"] > last_captured_os)]

    # Mantem apenas colunas necessarias
    df = df[
        [
            "data_index",
            "Início da Vigência da OS",
            "Arquivo OS",
            "Arquivo GTFS",
            "Link da OS",
            "Link do GTFS",
        ]
    ]

    log(f"Os info: {df.head()}")
    if len(df) >= 1:
        log("Nova OS encontrada!")
        data = df.to_dict(orient="records")[0]  # Converte o DataFrame para um dicionário
        flag_new_os = True  # Se houver mais de uma OS, é uma nova OS

        # converte "Início da Vigência da OS" de dd/mm/aaaa para aaaa-mm-dd
        data["Início da Vigência da OS"] = datetime.strptime(
            data["Início da Vigência da OS"], "%d/%m/%Y"
        ).strftime("%Y-%m-%d")

    return flag_new_os, data, data["data_index"], data["Início da Vigência da OS"]


@task(nout=2)
def get_raw_drive_files(os_control, local_filepath: list):
    """
    Downloads raw files from Google Drive and processes them.

    Args:
        os_control (dict): A dictionary containing information about the OS (Ordem de Serviço).
        local_filepath (list): A list of local file paths where the downloaded files will be saved.

    Returns:
        raw_filepaths (list): A list of file paths where the downloaded raw files are saved.
        primary_keys (list[list]): A list with the primary_keys for the tables.
    """

    raw_filepaths = []

    log(f"Baixando arquivos: {os_control}")

    # Autenticar usando o arquivo de credenciais
    credentials = service_account.Credentials.from_service_account_file(
        filename=os.environ["GOOGLE_APPLICATION_CREDENTIALS"],
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
    )

    # Criar o serviço da API Google Drive e Google Sheets
    drive_service = build("drive", "v3", credentials=credentials)

    # Baixa planilha de OS
    file_link = os_control["Link da OS"]
    file_bytes_os = download_xlsx(file_link=file_link, drive_service=drive_service)

    # Baixa GTFS
    file_link = os_control["Link do GTFS"]
    file_bytes_gtfs = download_file(file_link=file_link, drive_service=drive_service)

    # Salva os nomes das planilhas
    sheetnames = xl.load_workbook(file_bytes_os).sheetnames

    with zipfile.ZipFile(file_bytes_gtfs, "r") as zipped_file:
        for filename in list(constants.GTFS_TABLE_CAPTURE_PARAMS.value.keys()):
            if filename == "ordem_servico":

                processa_ordem_servico(
                    sheetnames=sheetnames,
                    file_bytes=file_bytes_os,
                    local_filepath=local_filepath,
                    raw_filepaths=raw_filepaths,
                )
            elif filename == "ordem_servico_trajeto_alternativo":

                processa_ordem_servico_trajeto_alternativo(
                    sheetnames=sheetnames,
                    file_bytes=file_bytes_os,
                    local_filepath=local_filepath,
                    raw_filepaths=raw_filepaths,
                )

            else:
                # Processa arquivos do GTFS
                data = zipped_file.read(filename + ".txt")
                data = data.decode(encoding="utf-8")

                # encontra a partição correta
                local_file_path = list(filter(lambda x: filename + "/" in x, local_filepath))[0]

                raw_file_path = save_raw_local_func(
                    data=data, filepath=local_file_path, filetype="txt"
                )
                log(f"Saved file: {raw_file_path}")

                raw_filepaths.append(raw_file_path)

    return raw_filepaths, list(constants.GTFS_TABLE_CAPTURE_PARAMS.value.values())
